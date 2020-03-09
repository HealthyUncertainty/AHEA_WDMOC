# -*- coding: utf-8 -*-
"""
The Sequencer is the master file for the model. It creates entities and runs them through the model based
on their current model state, characteristics, history, and underlying probabilistic distrubtions of the 
model parameters.

The sequencer will run "num_entities" entities through the model and output the resources and events experienced
by each entity as a series of lists.

@author: icromwell
"""

################################
# STEP 1 - LOAD IN SCENARIO PARAMETERS
cd "/Users/icromwell/Desktop/My Work/PhD Thesis/WDMOC/Chapter 6"
"""
import pickle
import numpy
import math
import random
import time
import copy
from openpyxl import load_workbook

from Glb_ApplyInit import ApplyInit
from Glb_CheckTime import CheckTime
from Glb_Estimates import Estimates
from Glb_Estimates import Estimate
from Chap6_SysP_ScreenAppt import ScreenApptScen
from SysP_ScreenAppt import ScreenAppt
from Chap6_NatHist_OralCancer_HPV import NatHistOCaHPV
from Chap6_NatHist_OralCancer import NatHistOCa
from Chap6_SysP_OPLmanage import OPLManage
from Chap6_SysP_IncidentCancer import IncidentCancerScen
from SysP_IncidentCancer import IncidentCancer
from SysP_Followup import Followup
from SysP_Terminal import Terminal

with open('estimates.pickle', 'rb') as f:
    estimates = pickle.load(f)
with open('regcoeffs.pickle', 'rb') as f:
    regcoeffs = pickle.load(f)
with open('costdict.pickle', 'rb') as f:
    CostDict = pickle.load(f)    
    
from Glb_CreateEntity import Entity    
from Chap6_SysP_Prevention import Chap6_Prevention
"""

##### Read in scenario-specific model parameter estimates
altbook = load_workbook('Chap6_Alt_Parameters.xlsx')
sheet = altbook["Inputs"]
alt_estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(alt_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(alt_estimates.Parameter)

estimates_orig = copy.deepcopy(estimates)

# Add scenario-specific costs to Cost Dictionary
def CDpop(param):
    return (param.type, param.mean, param.se)

CostDict['Experimental Surgery'] = CDpop(alt_estimates.Chap6_SurgCost)
CostDict['Experimental Chemo'] = CDpop(alt_estimates.Chap6_ChemoCost)
CostDict['Experimental Alcohol Cessation'] = CDpop(alt_estimates.Chap6_AlcCost)
CostDict['Experimental Smoking Cessation'] = CDpop(alt_estimates.Chap6_SmokeCost)
CostDict['Experimental Cancer Screening'] = CDpop(alt_estimates.Chap6_ScreenCost)
CostDict['Experimental HPV Vaccination'] = CDpop(alt_estimates.Chap6_VaxCost)

# Load file that has names/combinations of scenarios
scenbook = load_workbook('Scenario_Combo.xlsx')
sheet = scenbook["Sheet1"]
scenarios = []
for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    scenarios.append([cell.value for cell in line])
del(scenarios[0])

#############################################################
####### START SCENARIO LOOP #######

# Define the number of entities you want to model"
num_entities = 1200000
from Glb_AnalyzeOutput import Analyze_Output
output = Analyze_Output(estimates, CostDict)

# STEP 1 - DEFINE THE SCENARIO BEING ANALYZED
for scen in range(len(scenarios)):
    scen_num = scen + 1, "of", len(scenarios)
    Scenario_Prevention = scenarios[scen][1]
    Scenario_Screen = scenarios[scen][2]
    Scenario_Surg = scenarios[scen][3]
    Scenario_Chemo = scenarios[scen][4]
    Scenario_HPV = scenarios[scen][5]
    version = scenarios[scen][0]
    versionext = 'Chap6_' + version + '.csv'
    print("Prev:", Scenario_Prevention, "Screen:", Scenario_Screen, "Surg:", Scenario_Surg, "Chemo:", Scenario_Chemo, "HPV:", Scenario_HPV)
    
    resources = []
    events = []
    natHist = []
    QALY = []
    
    ################################
    # STEP 3 - RUN THE SEQUENCER
    
    print("NOW EVALUATING scenario", scen_num, ":", scenarios[scen][0])
    print(time.ctime(int(time.time())))
    
    # Load a pre-generated population of entities with necessary 'upstream' characteristics
    print("Load a population of entities", time.ctime(int(time.time())))
    looptime_start = time.time()
    loadname = 'Chap6_population'
    if Scenario_Prevention == 1:
        loadname += '_prev'
    if Scenario_Screen == 1:
        loadname += '_screen'
    if Scenario_HPV == 1:
        loadname += '_hpv'
    oplload = loadname + '_opl.pickle'
    nooplload = loadname + '_noopl.pickle'
    with open(oplload, 'rb') as f:
        oplents = pickle.load(f)
    with open(nooplload, 'rb') as f:
        nooplents = pickle.load(f)

    for k in range(0,2):
        Alt_Scenario = k
        print ("Model Arm", k, "@", round((time.time() - looptime_start)/60, 2))
        EntityList = []
        seq_naive = copy.deepcopy(oplents) + copy.deepcopy(nooplents)
        seq_infm = copy.deepcopy(oplents)
        if Alt_Scenario == 0:
            seqlist = seq_naive
        else:
            seqlist = seq_infm
        for ent in range(len(seqlist)):
            entity = seqlist[ent]
            entity.altscen = Alt_Scenario
            entity.Scenario_Prevention = Scenario_Prevention
            entity.Scenario_Screen = Scenario_Screen
            entity.Scenario_Surg = Scenario_Surg
            entity.Scenario_Chemo = Scenario_Chemo
            entity.Scenario_HPV = Scenario_HPV
            while True:
                CheckTime(entity, estimates, natHist, QALY)
                #People with a participating dentist undergo regular screening appointments    
                if entity.stateNum == 1.0:     
                    if Scenario_Screen == 1:
                        screenapptscen = ScreenApptScen(estimates, regcoeffs, alt_estimates)
                        screenapptscen.Process(entity)
                    else:
                        screenappt = ScreenAppt(estimates, regcoeffs)   
                        screenappt.Process(entity)

                #People with no dentist wait for disease event        
                if entity.stateNum == 1.8:
                    entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
                        
                #People with a detected premalignancy undergo regular follow-up        
                if entity.stateNum == 2.0:
                    entity.OPLflag = 1
                    oplmanage = OPLManage(estimates, regcoeffs)            
                    oplmanage.Process(entity) 
                
                #People with a detected cancer undergo treatment       
                if entity.stateNum == 3.0:
                    if Scenario_Surg == 1:
                        entity.RR_Surgery = alt_estimates.RR_Surgery.sample()
                    if Scenario_Chemo == 1:
                        entity.RR_Chemo = alt_estimates.RR_Chemo.sample()
                    if Scenario_HPV == 1:
                        incidentcancerscen = IncidentCancerScen(estimates, regcoeffs, alt_estimates)
                        incidentcancerscen.Process(entity)    
                    else:
                        incidentcancer = IncidentCancer(estimates, regcoeffs)
                        incidentcancer.Process(entity)
        
                #People who have been successfully treated undergo regular follow-up     
                if entity.stateNum == 4.0:
                    followup = Followup(estimates, regcoeffs)
                    followup.Process(entity)
           
                #People whose disease has entered remission after 10 years     
                if entity.stateNum == 4.8:
                    #entity is in remission, no further events occur
                    entity.allTime = entity.natHist_deathAge + 0.0001
                     
                #People with terminal disease receive palliative care     
                if entity.stateNum == 5.0:
                    terminal = Terminal(estimates, regcoeffs)
                    terminal.Process(entity)
              
                #The entity is dead      
                if entity.stateNum == 100:
                    entity.utility.append(('Dead', 0, entity.allTime))
                    if entity.death_type == 1:
                        entity.horizon_censor = 0
                    #print("Entity is", entity.death_desc, "at:", entity.time_death)
                    if entity.horizon_censor == 1:
                        events.append(('Entity reaches model time horizon', entity.allTime))
                        entity.death_type = 'Censored'
                    else:
                        events.append(('Entity dies', entity.allTime))
                    break
                
                # An error has occurred
                if entity.stateNum == 99:
                    print(i, "An error has occurred and the simulation must end")
                    print(entity.currentState)
                    break
                
            EntityList.append(entity)
            estimates = estimates_orig
        """
        Oversample = []    
        while len(Oversample) < 10000:
            entity = copy.deepcopy(random.choice(oplents))
            entity.altscen = Alt_Scenario
            while True:
                CheckTime(entity, estimates, natHist, QALY)
                #People with a participating dentist undergo regular screening appointments    
                if entity.stateNum == 1.0:     
                    if Scenario_Screen == 1:
                        screenapptscen = ScreenApptScen(estimates, regcoeffs, alt_estimates)
                        screenapptscen.Process(entity)
                    else:
                        screenappt = ScreenAppt(estimates, regcoeffs)   
                        screenappt.Process(entity)

                #People with no dentist wait for disease event        
                if entity.stateNum == 1.8:
                    entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
                        
                #People with a detected premalignancy undergo regular follow-up        
                if entity.stateNum == 2.0:
                    entity.OPLflag = 1
                    oplmanage = OPLManage(estimates, regcoeffs)            
                    oplmanage.Process(entity) 
                
                #People with a detected cancer undergo treatment       
                if entity.stateNum == 3.0:
                    if Scenario_Surg == 1:
                        entity.RR_Surgery = alt_estimates.RR_Surgery.sample()
                    if Scenario_Chemo == 1:
                        entity.RR_Chemo = alt_estimates.RR_Chemo.sample()
                    if Scenario_HPV == 1:
                        incidentcancerscen = IncidentCancerScen(estimates, regcoeffs, alt_estimates)
                        incidentcancerscen.Process(entity)    
                    else:
                        incidentcancer = IncidentCancer(estimates, regcoeffs)
                        incidentcancer.Process(entity)
        
                #People who have been successfully treated undergo regular follow-up     
                if entity.stateNum == 4.0:
                    followup = Followup(estimates, regcoeffs)
                    followup.Process(entity)
           
                #People whose disease has entered remission after 10 years     
                if entity.stateNum == 4.8:
                    #entity is in remission, no further events occur
                    entity.allTime = entity.natHist_deathAge + 0.0001
                     
                #People with terminal disease receive palliative care     
                if entity.stateNum == 5.0:
                    terminal = Terminal(estimates, regcoeffs)
                    terminal.Process(entity)
              
                #The entity is dead      
                if entity.stateNum == 100:
                    entity.utility.append(('Dead', 0, entity.allTime))
                    if entity.death_type == 1:
                        entity.horizon_censor = 0
                    #print("Entity is", entity.death_desc, "at:", entity.time_death)
                    if entity.horizon_censor == 1:
                        events.append(('Entity reaches model time horizon', entity.allTime))
                        entity.death_type = 'Censored'
                    else:
                        events.append(('Entity dies', entity.allTime))
                    break
                
                # An error has occurred
                if entity.stateNum == 99:
                    print(i, "An error has occurred and the simulation must end")
                    print(entity.currentState)
                    break
            
            if entity.OPLflag == 1:    
                Oversample.append(entity)
            estimates = estimates_orig
        """
        ################################
        # OPTIONAL STEP - SAVE OUTPUTS TO DISK
        if k == 0:
            AssayNaive = copy.deepcopy(EntityList)
            #Oversample_0 = copy.deepcopy(Oversample)
        else:
            AssayInformed = copy.deepcopy(EntityList)
            #Oversample_1 = copy.deepcopy(Oversample)
            
    now = time.time()
    seqtime = round((now - looptime_start)/60, 2)
    print("The sequencer simulated", num_entities, "entities. It took", seqtime, "minutes. You can do this.")

    # Estimate the LYG and QALY generated by the entities in the population
    NoOPL0_Surv = []
    NoOPL0_Cost = []
    NoOPL1_Surv = []
    NoOPL1_Cost = []
    print("Calculating LYG and QALY")
    for k in range(0,2):
        if k == 1:
            EList = AssayInformed
        elif k == 0:
            EList = AssayNaive
        NoOPL = []
        HasOPL = []
        for ent in range(len(EList)):
            if EList[ent].OPLflag == 0:
                NoOPL.append(EList[ent])
            else:
                HasOPL.append(EList[ent])
        if k == 0:
            NoOPL_0 = NoOPL
            HasOPL_0 = HasOPL
        elif k == 1:
            NoOPL_1 = NoOPL
            HasOPL_1 = HasOPL
            
    OPLSurv_0 = np.array([output.EntitySurvival(x) for x in HasOPL_0])
    OPLCost_0 = np.array([output.EntityCost(x) for x in HasOPL_0])
    OPLSurv_1 = np.array([output.EntitySurvival(x) for x in HasOPL_1])
    OPLCost_1 = np.array([output.EntityCost(x) for x in HasOPL_1])
    #OvrSurv_0 = np.array([output.EntitySurvival(x) for x in Oversample_0])
    #OvrCost_0 = np.array([output.EntityCost(x) for x in Oversample_0])
    #OvrSurv_1 = np.array([output.EntitySurvival(x) for x in Oversample_1])
    #OvrCost_1 = np.array([output.EntityCost(x) for x in Oversample_1])
    
    OPL0 = np.c_[OPLSurv_0, OPLCost_0, [1]*len(OPLSurv_0)]
    OPL1 = np.c_[OPLSurv_1, OPLCost_1, [1]*len(OPLSurv_1)]
    now = time.time()
    print("OPL cases done @ ", (now - looptime_start)/60, "minutes")
    
    #Ovr0 = np.c_[OvrSurv_0, OvrCost_0]
    #Ovr1 = np.c_[OvrSurv_1, OvrCost_1]
    #OversampleCEA = np.c_[Ovr0, Ovr1]
    #now = time.time()
    #print("Oversampling done @ ", (now - looptime_start)/60, "minutes")
    
    discountrate = estimates.DiscountRate.mean
    disc_rate = 1 - (1 - discountrate)**(1 / 365)
    maxdays = estimates.timehorizon.mean*365
    day = 0
    maxLYG = 0
    while day < maxdays:
        maxLYG += (1/365)*(1+disc_rate)**(-day)
        day+=1
    
    for i in range(len(NoOPL_0)):
        if NoOPL_0[i].time_death == maxdays:
            NoOPL0_Surv.append([maxLYG, maxLYG*estimates.Util_Well.sample()])
        else:
            NoOPL0_Surv.append(output.EntitySurvival(NoOPL_0[i]))
        NoOPL0_Cost.append(output.EntityCost(NoOPL_0[i]))
    
    now = time.time()
    print("No OPL, Assay Naive done @", (now - looptime_start)/60, "minutes")
        
    NoOPL0 = np.c_[np.array(NoOPL0_Surv)[:,0], np.array(NoOPL0_Surv)[:,1], NoOPL0_Cost, [0]*len(NoOPL0_Surv)]
    Output_0 = np.vstack((NoOPL0, OPL0))
    Output_1 = np.vstack((NoOPL0, OPL1))
    
    if np.size(Output_1,0) > np.size(Output_0,0):
        Output_1 = numpy.delete(Output_1, slice(0, (np.size(Output_1,0) - np.size(Output_0,0))), axis=0)
        OutputCEA = np.c_[Output_0, Output_1]
    elif np.size(Output_1,0) < np.size(Output_0,0):
        Output_0 = numpy.delete(Output_0, slice(0, (np.size(Output_0,0) - np.size(Output_1,0))), axis=0)
        OutputCEA = np.c_[Output_0, Output_1]
    else:
        OutputCEA = np.c_[Output_0, Output_1]
    
    # Output results as csv
    print("Saving...")
    numpy.savetxt(versionext, OutputCEA, delimiter=",")
    
    #overext = 'Chap6_Oversample' + version + '.csv'
    #numpy.savetxt(overext, OversampleCEA, delimiter=",")
    now = time.time()
    print(scenarios[scen][0], "finished. This process took ", (now - looptime_start)/60, "minutes")
    ####################################################
    # FOOTNOTE:
    #
    #   1 - The clock moves forward an arbitrary number of days, but is reset to the next natural
    #       history or disease event by 'Glb_Checktime.py'. The purpose of moving the clock forward is
    #       simply to prompt advancement to the next event.
    
Output = [AssayNaive, AssayInformed]
with open('Chap6_Entities_Baseline.pickle', 'wb') as inputs:
    pickle.dump(Output, inputs, pickle.HIGHEST_PROTOCOL)


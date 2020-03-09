# -*- coding: utf-8 -*-
"""
The Sequencer is the master file for the model. It creates entities and runs them through the model based
on their current model state, characteristics, history, and underlying probabilistic distrubtions of the 
model parameters.

The sequencer will run "num_entities" entities through the model and output the resources and events experienced
by each entity as a series of lists.

@author: icromwell
"""
############################################################################################
############################################################################################
# LOAD SOME NECESSARY PACKAGES AND FUNCTIONS

import time
import math
import pickle
import random
from openpyxl import load_workbook                  # Load the import function
import numpy

# Import Parameter Estimates from the table"

from Glb_Estimates import Estimates
from Glb_Estimates import Estimate

inbook = load_workbook('InputParameters.xlsx')
sheet = inbook["Inputs"]
estimates = Estimates()
for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(estimates.Parameter)
with open('estimates.pickle', 'wb') as inputs:
    pickle.dump(estimates, inputs, pickle.HIGHEST_PROTOCOL)

# Import Regression Coefficients from the table

regsheet = inbook["RegCoeffs"]
source = []                                           # A list to hold data

"Convert the openpyxl object into a useable form"

for row in list(regsheet.rows)[1:]:
    args = [cell.value for cell in row]
    source.append(args)
    
for row in range(len(source)):
    source[row][0] = str(source[row][0])
    source[row][1] = str(source[row][1])

"Create a multi-level dictionary to hold each parameter from the regression model:"
    
config = {}         # creates the blank dictionary
for param, factor, vartype, level, mean, SE in source:
    SE = SE if SE else 0    # If SE is blank, enter zero
    vartype = vartype if vartype else 0
    mean = mean if mean not in ("ref", None) else 0     # Reference category = 0
    if param not in config:
        config[param] = {}
    
    if level:
        if factor not in config[param]:
            config[param][factor] = {"vartype": vartype}
        config[param][factor][level] = {"mean": mean, "SE": SE}
    else:
        config[param][factor] = {"vartype": vartype, "mean": mean, "SE": SE}

with open('regcoeffs.pickle', 'wb') as regcoeffs:
    pickle.dump(config, regcoeffs, pickle.HIGHEST_PROTOCOL)
    
# Import cost estimates from the table

costsheet = inbook["Costs"]
cost_estimates = Estimates()
for line in costsheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(cost_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(cost_estimates.Parameter)

# Create a dictionary of unit costs that the program will read from
CostDict = {}

for i in range(0, costsheet.max_row):
    cost_name = str(costsheet.cell(row = i+1, column = 1).value)
    cost_type = costsheet.cell(row = i+1, column = 2).value
    cost_mean = costsheet.cell(row = i+1, column = 3).value
    cost_se = costsheet.cell(row = i+1, column = 4).value
    CostDict[cost_name] = (cost_type, cost_mean, cost_se)
del(CostDict['Parameter'])
#del(CostDict['None'])
with open('costdict.pickle', 'wb') as costdict:
    pickle.dump(CostDict, costdict, pickle.HIGHEST_PROTOCOL)

with open('estimates.pickle', 'rb') as f:
    estimates = pickle.load(f)
with open('regcoeffs.pickle', 'rb') as f:
    regcoeffs = pickle.load(f)
with open('costdict.pickle', 'rb') as f:
    CostDict = pickle.load(f)

#############################################################################################
############################################################################################

################################
# STEP 1 - LOAD IN SCENARIO PARAMETERS

# Alternative parameters are the values used for the specific scenarios considered in this
# manuscript (New Drug, Improved Screening, etc.) and are adjunct to the core WDMOC parameters
# stored in 'InputParameters.xlsx'

altbook = load_workbook('Alt_Parameters.xlsx')
sheet = altbook["Inputs"]
alt_estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(alt_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(alt_estimates.Parameter)

estimates_orig = estimates

def CDpop(param):
    return (param.type, param.mean, param.se)

CostDict['Experimental Surgery'] = CDpop(alt_estimates.Alt_SurgCost)
CostDict['Experimental Chemo'] = CDpop(alt_estimates.Alt_ChemoCost)
CostDict['Experimental Alcohol Cessation'] = CDpop(alt_estimates.Alt_AlcCost)
CostDict['Experimental Smoking Cessation'] = CDpop(alt_estimates.Alt_SmokeCost)
CostDict['Experimental Cancer Screening'] = CDpop(alt_estimates.Alt_ScreenCost)
CostDict['Experimental HPV Vaccination'] = CDpop(alt_estimates.Alt_VaxCost)

################################
# STEP 2 - DEFINE THE SCENARIO BEING ANALYZED
# 1 - include the effects of this policy scenario; 0 - do not include
Scenario_Prevention = 0
Scenario_Screen = 0
Scenario_Surg = 1
Scenario_Chemo = 0
Scenario_HPV = 0

################################
# STEP 3 - RUN THE SEQUENCER
"Define the number of entities you want to model"
num_entities = 100000

import time
looptime_start = time.time()
for k in range(0,2):
    Alt_Scenario = k
    EntityList = []
    for i in range(0, num_entities):
        entity_num = 'entity' + str(i + 1)                  # Create an entity for each iteration of the model
        if i % 50000 == 0:
            print("Entity ", i, "of scenario ", k)
        
        from Glb_CreateEntity import Entity
        entity_num = Entity()
        entity = entity_num                             # Identify the current entity as the most recently created one
        
        entity.Scenario_Prev = Scenario_Prevention
        entity.Scenario_Screen = Scenario_Screen
        entity.Scenario_Surg = Scenario_Surg
        entity.Scenario_Chemo = Scenario_Chemo
        entity.Scenario_HPV = Scenario_HPV
        entity.scenario_desc = []
        
        #print("Entity %2.0f is created"%(i+1))

        resources = []
        events = []
        natHist = []
        QALY = []
    
        while True:
            #Apply Demographic Characteristics and Natural History to a newly-created entity
            if entity.stateNum == 0.0:
                from Glb_ApplyInit import ApplyInit
                applyinit = ApplyInit(estimates)
                applyinit.Process(entity)
                if Scenario_Prevention == 1:
                    from SysP_Prevention import Prevention
                    prevention = Prevention(alt_estimates)
                    prevention.Process(entity)
                # HPV vaccination affects natural history of OPL progression to cancer
                if Scenario_HPV == 1:
                    alt_regcoeffs = regcoeffs
                    from Alt_NatHist_OralCancer_HPV import NatHistOCa
                    nathistoca = NatHistOCa(estimates, alt_regcoeffs, alt_estimates)
                else:
                    from Alt_NatHist_OralCancer import NatHistOCa
                    nathistoca = NatHistOCa(estimates, regcoeffs)
                nathistoca.Process(entity, natHist)

                if entity.hasOPL == 1: 
                    entity.OPLflag = 1
                else: 
                    entity.OPLflag = 0
            
            #PROBABILITY NODE: Does this person regularly see a dentist?
                # If yes - develop OPL while undergoing regular observations (state 1.0)
                # If no - develop OPL and possibly cancer (state 1.8)
            
            # In this scenario, all entities start at OPL management    
            if entity.stateNum == 0.1:
                entity.altscen = Alt_Scenario
                if entity.hasDentist == 1:
                    entity.stateNum = 1.0
                    entity.currentState = "1.0 - Start regular dental screening"
                elif entity.hasDentist == 0:
                    entity.stateNum = 1.8
                    entity.currentState = "1.8 - No access to dentist"
                
            ### Advance the clock to next scheduled event (NatHist, Sysp, Recurrence, Death) ###
            
            from Glb_CheckTime import CheckTime
            CheckTime(entity, estimates, natHist, QALY)
            
            ### Run next scheduled event/process according to state ###
            
            #People with a participating dentist undergo regular screening appointments    
            if entity.stateNum == 1.0:     
                if Scenario_Screen == 1:
                    from Alt_SysP_ScreenAppt import ScreenAppt
                    screenappt = ScreenAppt(estimates, regcoeffs, alt_estimates)   
                else:
                    from SysP_ScreenAppt import ScreenAppt
                    screenappt = ScreenAppt(estimates, regcoeffs)            
                screenappt.Process(entity)
               
            #People with no dentist wait for disease event        
            if entity.stateNum == 1.8:
                entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
                    
            #People with a detected premalignancy undergo regular follow-up        
            if entity.stateNum == 2.0:
                from SysP_OPLmanage import OPLManage
                oplmanage = OPLManage(estimates, regcoeffs)            
                oplmanage.Process(entity) 
            
            #People with a detected cancer undergo treatment       
            if entity.stateNum == 3.0:
                if Scenario_Surg == 1:
                    entity.RR_Surgery = alt_estimates.RR_Surgery.sample()
                if Scenario_Chemo == 1:
                    entity.RR_Chemo = alt_estimates.RR_Chemo.sample()
                if Scenario_HPV == 1:
                    from Alt_SysP_IncidentCancer import IncidentCancer
                    incidentcancer = IncidentCancer(estimates, regcoeffs, alt_estimates)
                else:
                    from SysP_IncidentCancer import IncidentCancer
                    incidentcancer = IncidentCancer(estimates, regcoeffs)
                incidentcancer.Process(entity)        
    
            #People who have been successfully treated undergo regular follow-up     
            if entity.stateNum == 4.0:
                from SysP_Followup import Followup
                followup = Followup(estimates, regcoeffs)
                followup.Process(entity)
       
            #People whose disease has entered remission after 10 years     
            if entity.stateNum == 4.8:
                #entity is in remission, no further events occur
                entity.allTime = entity.natHist_deathAge + 0.0001
                 
            #People with terminal disease receive palliative care     
            if entity.stateNum == 5.0:
                from SysP_Terminal import Terminal
                terminal = Terminal(estimates, regcoeffs)
                terminal.Process(entity)
          
            #The entity is dead      
            if entity.stateNum == 100:
                entity.utility.append(('Dead', 0, entity.allTime))
                if entity.death_type == 1:
                    entity.horizon_censor = 0
                #print("Entity is", entity.death_desc, "at:", entity.time_death)
                if entity.horizon_censor == 1:
                    events.append(('Entity reaches model time horizon', entity.allTime))
                    entity.death_type = 'Censored'
                else:
                    events.append(('Entity dies', entity.allTime))
                break
            
            # An error has occurred
            if entity.stateNum == 99:
                print(i, "An error has occurred and the simulation must end")
                print(entity.currentState)
                break
            
        EntityList.append(entity)
        estimates = estimates_orig
    
    ################################
    # OPTIONAL STEP - SAVE OUTPUTS TO DISK
    if Alt_Scenario == 0:
        AssayNaive = EntityList
    elif Alt_Scenario == 1:
        AssayInformed = EntityList

now = time.time()
seqtime = round((now - looptime_start)/60, 2)
print("The sequencer simulated", num_entities, "entities. It took", seqtime, "minutes.")

# Estimate the LYG and QALY generated by the entities in the population
print("Calculating LYG and QALY")
for k in range(0,2):
    if k == 1:
        EList = AssayInformed
    elif k == 0:
        EList = AssayNaive
    NoOPL = []
    HasOPL = []
    for ent in range(len(EList)):
        if EList[ent].OPLflag == 0:
            NoOPL.append(EList[ent])
        else:
            HasOPL.append(EList[ent])
    if k == 0:
        NoOPL_0 = NoOPL
        HasOPL_0 = HasOPL
    elif k == 1:
        NoOPL_1 = NoOPL
        HasOPL_1 = HasOPL

from Glb_AnalyzeOutput import Analyze_Output
output = Analyze_Output(estimates, CostDict)
        
OPLSurv_0 = np.array([output.EntitySurvival(x) for x in HasOPL_0])
OPLCost_0 = np.array([output.EntityCost(x) for x in HasOPL_0])
OPLSurv_1 = np.array([output.EntitySurvival(x) for x in HasOPL_1])
OPLCost_1 = np.array([output.EntityCost(x) for x in HasOPL_1])
OPL0 = np.c_[OPLSurv_0, OPLCost_0, [1]*len(OPLSurv_0)]
OPL1 = np.c_[OPLSurv_1, OPLCost_1, [1]*len(OPLSurv_1)]

now = time.time()
print("OPL cases done @ ", (now - looptime_start)/60, "minutes")

discountrate = estimates.DiscountRate.mean
disc_rate = 1 - (1 - discountrate)**(1 / 365)
maxdays = estimates.timehorizon.mean*365
day = 0
maxLYG = 0
while day < maxdays:
    maxLYG += (1/365)*(1+disc_rate)**(-day)
    day+=1

NoOPL0_Surv = []
NoOPL0_Cost = []
NoOPL1_Surv = []
NoOPL1_Cost = []
for i in range(len(NoOPL_0)):
    if NoOPL_0[i].time_death == maxdays:
        NoOPL0_Surv.append([maxLYG, maxLYG*estimates.Util_Well.sample()])
    else:
        NoOPL0_Surv.append(output.EntitySurvival(NoOPL_0[i]))
    NoOPL0_Cost.append(output.EntityCost(NoOPL_0[i]))

now = time.time()
print("No OPL, Assay Naive done @", (now - looptime_start)/60, "minutes")
    
for i in range(len(NoOPL_1)):
    if NoOPL_1[i].time_death == maxdays:
        NoOPL1_Surv.append([maxLYG, maxLYG*estimates.Util_Well.sample()])
    else:
        NoOPL1_Surv.append(output.EntitySurvival(NoOPL_1[i]))
    NoOPL1_Cost.append(output.EntityCost(NoOPL_1[i]))
now = time.time()
print("No OPL, Assay Informed done @", (now - looptime_start)/60, "minutes")


NoOPL0 = np.c_[np.array(NoOPL0_Surv)[:,0], np.array(NoOPL0_Surv)[:,1], NoOPL0_Cost, [0]*len(NoOPL0_Surv)]
NoOPL1 = np.c_[np.array(NoOPL1_Surv)[:,0], np.array(NoOPL1_Surv)[:,1], NoOPL1_Cost, [0]*len(NoOPL1_Surv)]

Output_0 = np.vstack((OPL0, NoOPL0))
Output_1 = np.vstack((OPL1, NoOPL1))

OutputCEA = np.c_[Output_0, Output_1]

# Output results as csv
# This step allows you to name the outputs so you can keep track of which file you want to analyze
versionext = 'Scenario_Output.csv'
print("Saving...")
numpy.savetxt(versionext, OutputCEA, delimiter=",")
now = time.time()
print("Done: this process took ", (now - looptime_start)/60, "minutes")
print("Prev:", Scenario_Prevention, "Screen:", Scenario_Screen, "Surg:", Scenario_Surg, "Chemo:", Scenario_Chemo, "HPV:", Scenario_HPV)


####################################################
# FOOTNOTE:
#
#   1 - The clock moves forward an arbitrary number of days, but is reset to the next natural
#       history or disease event by 'Glb_Checktime.py'. The purpose of moving the clock forward is
#       simply to prompt advancement to the next event.

# -*- coding: utf-8 -*-
"""
A program to read and apply costs from a completed simulation. 

The program reads in unit costs from an Excel file and applies them to each instance
of the resource from the ResourceList array output from the Sequencer. It returns a
list of discounted costs


@author: icromwell
"""

############################################################################################
# Load some necessary packages and functions

from openpyxl import load_workbook                  # Load the import function
import numpy
import random
import simpy

# Import Parameter Estimates from the table"

from Import_Varnames import Estimates
from Import_Varnames import Estimate

inbook = load_workbook('InputParameters.xlsx')
sheet = inbook["Costs"]
estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))

del(estimates.Parameter)

with open('estimates.pickle', 'wb') as inputs:
    pickle.dump(estimates, inputs, pickle.HIGHEST_PROTOCOL)
    

# Import Regression Coefficients from the table

regsheet = inbook["RegCoeffs"]
source = []                                           # A list to hold data

"Convert the openpyxl object into a useable form"

for row in regsheet.rows[1:]:
    args = [cell.value for cell in row]
    source.append(args)
    
for row in range(len(source)):
    source[row][0] = str(source[row][0])
    source[row][1] = str(source[row][1])

"Create a multi-level dictionary to hold each parameter from the regression model:"
    
config = {}         # creates the blank dictionary
for param, factor, vartype, level, mean, SE in source:
    SE = SE if SE else 0    # If SE is blank, enter zero
    vartype = vartype if vartype else 0
    mean = mean if mean not in ("ref", None) else 0     # Reference category = 0
    if param not in config:
        config[param] = {}
    
    if level:
        if factor not in config[param]:
            config[param][factor] = {"vartype": vartype}
        config[param][factor][level] = {"mean": mean, "SE": SE}
    else:
        config[param][factor] = {"vartype": vartype, "mean": mean, "SE": SE}

with open('regcoeffs.pickle', 'wb') as regcoeffs:
    pickle.dump(config, regcoeffs, pickle.HIGHEST_PROTOCOL)
############################################################################################     

# Load in the resource utilization list from the model run
ResourceList = numpy.load('ResourceList.npy')

# A function to randomly sample cost inputs from the dictionary according to a Gamma distribution

class GenCost:
    
    def __init__(self, estimates, regcoeffs, resources):
        self._estimates = estimates
        self._regcoeffs = regcoeffs
        self._resources = ResourceList

    def CostMean(unit, costdict):
    
        x = costdict[unit][0]
        y = costdict[unit][1]
        gdist_shape = x**2/y**2                                    # A formula to produce the shape parameter
        gdist_scale = y**2/x                                       # A formula to produce the scale parameter
        samp_value = numpy.random.gamma(gdist_shape, gdist_scale)
        return samp_value
        
    def readVal(self, entity, param):
        
        # Is the parameter being estimated contained within the Excel sheet?
        if param in self._regcoeffs:
   
            # The sum of the coefficients starts at zero
            coeff = 0
     
            # For a given factor of a parameter within the Excel sheet
            for factor in self._regcoeffs[param].keys():      
                
                # Identify the intercept
                if factor == 'Intercept':            
                    Intercept = self._regcoeffs[param]['Intercept']['mean']
                    
                # Identify the shape parameter from the output                 
                elif factor == 'Sigma':                
                    Sigma = self._regcoeffs[param]['Sigma']['mean']
                
                # Identify values for all other coefficients
                elif factor in entity.__dict__.keys():   
                    value = getattr(entity, factor)
                    
                    if self._regcoeffs[param][factor]['vartype'] == 2:
                        coeff += self._regcoeffs[param][factor]['mean'] * value
                    else:
                        coeff += self._regcoeffs[param][factor][value]['mean']
            
                # If the entity doesn't have the required factor    
                elif factor not in entity.__dict__.keys():                                  
                    entity.stateNum = 99
                    entity.currentState = "Error - could not estimate %s as entity is missing %s"%(param, factor)
                
            # Produce an estimate of time from the regression
            mu = Intercept + coeff         
            shape = Sigma
            scale = math.exp(mu)/Sigma
            
            self.mu = mu
            self.shape = shape
            self.scale = scale
            
        else:
            entity.stateNum = 99
            entity.currentState = "Error - tried to produce a cost estimate for a variable that does not exist"
            
    # Randomly sample an event time for the entity from a Weibull distribution            
    def estCost(self):                 
        estimate_cost = numpy.random.gamma(self.shape, self.scale)
        return estimate_cost

from Glb_RegGenCost import RegGenCost

def 

reggencost = RegGenCost(estimates, regcoeffs)
cost_primtx = RegGenCost.readVal(entity, 'Cost_TxPrim')

# Create a dictionary of unit costs that the program will read from

CostDict = {}

maxrow = sheet.get_highest_row()

for i in range(maxrow):
    
    cost_name = str(sheet.cell(row = i+1, column = 0).value)
    cost_mean = sheet.cell(row = i+1, column = 2).value
    cost_se = sheet.cell(row = i+1, column = 3).value
    CostDict[cost_name] = (cost_mean, cost_se)

del(CostDict['None'])


# Create a list with costs generated by each entity

sumCost = []

for j in range(ResourceList.shape[0]):

    costList = []

    if ResourceList[j] == []:                                   # If no costs have been generated
        costList.append(0)

    else:    
        for i in range(int(ResourceList.shape[0])):
    
            unit = ResourceList[j][i][0]
            year = (float(ResourceList[j][i][1])/365.25) - 1     # Years elapsed for the discounting function
            
            disc_rate = 1 / (1 + CostDict['Discount'][0])**year              # Apply a discount rate for future costs
            
            costList.append(GenCost(unit, CostDict)*disc_rate)
            
    sumCost.append(sum(costList))

cd "/Users/icromwell/Desktop/My Work/PhD Thesis/WDMOC/Chapter 6"

############################################################################################
# Load some necessary packages and functions

import time
import pickle
import pandas
from openpyxl import load_workbook                  # Load the import function
import numpy
from collections import Counter

# Import Parameter Estimates from the table"

from Glb_Estimates import Estimates
from Glb_Estimates import Estimate

inbook = load_workbook('Chap6_InputParameters.xlsx')
sheet = inbook["Inputs"]
estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))

del(estimates.Parameter)

with open('estimates.pickle', 'wb') as inputs:
    pickle.dump(estimates, inputs, pickle.HIGHEST_PROTOCOL)
    
# Import Regression Coefficients from the table

regsheet = inbook["RegCoeffs"]
source = []                                           # A list to hold data

"Convert the openpyxl object into a useable form"

for row in list(regsheet.rows)[1:]:
    args = [cell.value for cell in row]
    source.append(args)
    
for row in range(len(source)):
    source[row][0] = str(source[row][0])
    source[row][1] = str(source[row][1])

"Create a multi-level dictionary to hold each parameter from the regression model:"
    
config = {}         # creates the blank dictionary
for param, factor, vartype, level, mean, SE in source:
    SE = SE if SE else 0    # If SE is blank, enter zero
    vartype = vartype if vartype else 0
    mean = mean if mean not in ("ref", None) else 0     # Reference category = 0
    if param not in config:
        config[param] = {}
    
    if level:
        if factor not in config[param]:
            config[param][factor] = {"vartype": vartype}
        config[param][factor][level] = {"mean": mean, "SE": SE}
    else:
        config[param][factor] = {"vartype": vartype, "mean": mean, "SE": SE}

with open('regcoeffs.pickle', 'wb') as regcoeffs:
    pickle.dump(config, regcoeffs, pickle.HIGHEST_PROTOCOL)

# Import cost estimates from the table

costsheet = inbook["Costs"]
cost_estimates = Estimates()
for line in costsheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(cost_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(cost_estimates.Parameter)

# Create a dictionary of unit costs that the program will read from
CostDict = {}

for i in range(0, costsheet.max_row):
    cost_name = str(costsheet.cell(row = i+1, column = 1).value)
    cost_type = costsheet.cell(row = i+1, column = 2).value
    cost_mean = costsheet.cell(row = i+1, column = 3).value
    cost_se = costsheet.cell(row = i+1, column = 4).value
    CostDict[cost_name] = (cost_type, cost_mean, cost_se)
del(CostDict['Parameter'])
#del(CostDict['None'])

# A program to produce mean estimates of Cost, LYG, and QALY
from Glb_AnalyzeOutput import Analyze_Output

# A function to extract data from a list of entities
def genVarList(entityrecord, var):
        retvar = []
        for i in range(0, len(entityrecord)):
            if hasattr(entityrecord[i], var) == True:
                retvar.append(getattr(entityrecord[i], var))
        return retvar
        
# A function to calculate mean cost from an entity
def CostEst(unit, costdict):
    # Fee-for-service and other fixed-value resources have a unit cost equal to the mean
    if costdict[unit][0] == 1:
        samp_value = costdict[unit][1]
    
    # Other kinds of costs are sampled from a gamma distribution based on their mean and sd
    elif costdict[unit][0] == 2:
        x = costdict[unit][0]
        y = costdict[unit][1]
        gdist_shape = x**2/y**2                                    # A formula to produce the shape parameter
        gdist_scale = y**2/x                                       # A formula to produce the scale parameter
        samp_value = numpy.random.gamma(gdist_shape, gdist_scale)
        
    # Return an error if no variable type is specified
    else:
        print("Please specify a variable TypeNo for", unit, "in the parameter table")
        
    return samp_value

#############################################################################################

with open('estimates.pickle', 'rb') as f:
    estimates = pickle.load(f)

with open('regcoeffs.pickle', 'rb') as f:
    regcoeffs = pickle.load(f)

# Load SA values from Excel sheet

SAbook = load_workbook('Chap6_Parameters.xlsx')
SAsheet = SAbook["Sheet1"]
SAVals = []

for line in SAsheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    if not line[1].value:
        continue
    SAVals.append((line[0].value, line[1].value, line[3].value, line[4].value))
del(SAVals[0])

from Glb_AnalyzeOutput import Analyze_Output
output = Analyze_Output(estimates, CostDict)

"Define the number of entities you want to model in each iteration"
num_entities = 1000

test = []
# Run the sequencer for each sensitivity parameter
for inputval, baseline, hival, loval in SAVals:
    estimates.__dict__[inputval].mean = baseline
    print("Now evaluating", inputval)
    val = hival
    for j in range(0,2):
        estimates.__dict__[inputval].mean = val
        regcoeffs = regcoeffs
        param = inputval
        
        ResourceList = []
        EventsList = []
        CostList = []
        SurvivalList = []
        IterList = []
        
        "Run the sequencer"
        for k in range(0,2):
            Alt_Scenario = k
            EntityList = []
            for i in range(0, num_entities):
                entity_num = 'entity' + str(i + 1)                  # Create an entity for each iteration of the model
                
                from Glb_CreateEntity import Entity
                entity_num = Entity()
                entity = entity_num                             # Identify the current entity as the most recently created one
                #print("Entity %2.0f is created"%(i+1))
        
                resources = []
                events = []
                natHist = []
                QALY = []
            
                while True:
                    #Apply Demographic Characteristics and Natural History to a newly-created entity
                    if entity.stateNum == 0.0:
                        from Glb_ApplyInit import ApplyInit
                        applyinit = ApplyInit(estimates)
                        applyinit.Process(entity)
                        #entity.probOPL = 1      # All entities start with OPL
                        from Chap6_NatHist_OralCancer import NatHistOCa
                        nathistoca = NatHistOCa(estimates, regcoeffs)
                        nathistoca.Process(entity, natHist)
                    
                    #PROBABILITY NODE: Does this person regularly see a dentist?
                        # If yes - develop OPL while undergoing regular observations (state 1.0)
                        # If no - develop OPL and possibly cancer (state 1.8)
                    
                    # In this scenario, all entities start at OPL management    
                    if entity.stateNum == 0.1:
                        entity.altscen = Alt_Scenario
                        if entity.hasDentist == 1:
                            entity.stateNum = 1.0
                            entity.currentState = "1.0 - Start regular dental screening"
                        elif entity.hasDentist == 0:
                            entity.stateNum = 1.8
                            entity.currentState = "1.8 - No access to dentist"
                        #else:
                            #print("The entity has not been assigned a value for 'hasDentist'. The simulation must terminate")
                            #entity.stateNum = 0.9
                        
                    ### Advance the clock to next scheduled event (NatHist, Sysp, Recurrence, Death) ###
                    
                    from Glb_CheckTime import CheckTime
                    CheckTime(entity, estimates, natHist, QALY)
                    
                    ### Run next scheduled event/process according to state ###
                    
                    #People with a participating dentist undergo regular screening appointments    
                    if entity.stateNum == 1.0:                
                        from SysP_ScreenAppt import ScreenAppt
                        screenappt = ScreenAppt(estimates, regcoeffs)            
                        screenappt.Process(entity)
                       
                    #People with no dentist wait for disease event        
                    if entity.stateNum == 1.8:
                        entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
                            
                    #People with a detected premalignancy undergo regular follow-up        
                    if entity.stateNum == 2.0:
                        from Chap6_SysP_OPLmanage import OPLManage
                        oplmanage = OPLManage(estimates, regcoeffs)            
                        oplmanage.Process(entity) 
                    
                    #People with a detected cancer undergo treatment       
                    if entity.stateNum == 3.0:
                        from SysP_IncidentCancer import IncidentCancer
                        incidentcancer = IncidentCancer(estimates, regcoeffs)
                        incidentcancer.Process(entity)        
            
                    #People who have been successfully treated undergo regular follow-up     
                    if entity.stateNum == 4.0:
                        from SysP_Followup import Followup
                        followup = Followup(estimates, regcoeffs)
                        followup.Process(entity)
               
                    #People whose disease has entered remission after 10 years     
                    if entity.stateNum == 4.8:
                        #entity is in remission, no further events occur
                        entity.allTime = entity.natHist_deathAge + 0.0001
                         
                    #People with terminal disease receive palliative care     
                    if entity.stateNum == 5.0:
                        from SysP_Terminal import Terminal
                        terminal = Terminal(estimates, regcoeffs)
                        terminal.Process(entity)
                  
                    #The entity is dead      
                    if entity.stateNum == 100:
                        entity.utility.append(('Dead', 0, entity.allTime))
                        if entity.death_type == 1:
                            entity.horizon_censor = 0
                        #print("Entity is", entity.death_desc, "at:", entity.time_death)
                        if entity.horizon_censor == 1:
                            events.append(('Entity reaches model time horizon', entity.allTime))
                            entity.death_type = 'Censored'
                        else:
                            events.append(('Entity dies', entity.allTime))
                        break
                    
                    # An error has occurred
                    if entity.stateNum == 99:
                        print(i, "An error has occurred and the simulation must end")
                        print(entity.currentState)
                        break
                    
                EntityList.append(entity)
    
            # Estimate the costs generated by the entities in the population
            CohortCost = []
            for i in range(len(EntityList)):
                entity = EntityList[i]
                CohortCost.append(output.EntityCost(entity))
                
            # Estimate the LYG and QALY generated by the entities in the population
            CohortSurvival = np.array([output.EntitySurvival(x) for x in EntityList])
            CohortLYG = CohortSurvival[:,0]
            CohortQALY = CohortSurvival[:,1]
            OPLStatus = np.array([x.OPLStatus for x in EntityList])
            CohortCEA = np.c_[CohortSurvival, np.array(CohortCost), OPLStatus]
            
            ################################
            # OPTIONAL STEP - SAVE OUTPUTS TO DISK
            if Alt_Scenario == 0:
                AssayNaive = CohortCEA
            elif Alt_Scenario == 1:
                AssayInformed = CohortCEA
        test.append(val)
        val = loval
    
    # Join outputs
    OutputCEA = np.c_[AssayNaive, AssayInformed]
    
    # Output results as csv
    versionext = 'Chap6_Output - ' + param + '.csv'
    numpy.savetxt(versionext, OutputCEA, delimiter=",")

    estimates.__dict__[inputval].mean = baseline



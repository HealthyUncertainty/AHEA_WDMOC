############################################################################################
############################################################################################
# LOAD SOME NECESSARY PACKAGES AND FUNCTIONS

import time
import pickle
import random
from openpyxl import load_workbook                  # Load the import function
import numpy

# Import Parameter Estimates from the table"

from Glb_Estimates import Estimates
from Glb_Estimates import Estimate

inbook = load_workbook('Chap5_InputParameters.xlsx')
sheet = inbook["Inputs"]
estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))

del(estimates.Parameter)

with open('estimates.pickle', 'wb') as inputs:
    pickle.dump(estimates, inputs, pickle.HIGHEST_PROTOCOL)

# Import Regression Coefficients from the table

regsheet = inbook["RegCoeffs"]
source = []                                           # A list to hold data

"Convert the openpyxl object into a useable form"

for row in list(regsheet.rows)[1:]:
    args = [cell.value for cell in row]
    source.append(args)
    
for row in range(len(source)):
    source[row][0] = str(source[row][0])
    source[row][1] = str(source[row][1])

"Create a multi-level dictionary to hold each parameter from the regression model:"
    
config = {}         # creates the blank dictionary
for param, factor, vartype, level, mean, SE in source:
    SE = SE if SE else 0    # If SE is blank, enter zero
    vartype = vartype if vartype else 0
    mean = mean if mean not in ("ref", None) else 0     # Reference category = 0
    if param not in config:
        config[param] = {}
    
    if level:
        if factor not in config[param]:
            config[param][factor] = {"vartype": vartype}
        config[param][factor][level] = {"mean": mean, "SE": SE}
    else:
        config[param][factor] = {"vartype": vartype, "mean": mean, "SE": SE}

with open('regcoeffs.pickle', 'wb') as regcoeffs:
    pickle.dump(config, regcoeffs, pickle.HIGHEST_PROTOCOL)
    
# Import cost estimates from the table

costsheet = inbook["Costs"]
cost_estimates = Estimates()
for line in costsheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(cost_estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))
del(cost_estimates.Parameter)

# Create a dictionary of unit costs that the program will read from
CostDict = {}

for i in range(0, costsheet.max_row):
    cost_name = str(costsheet.cell(row = i+1, column = 1).value)
    cost_type = costsheet.cell(row = i+1, column = 2).value
    cost_mean = costsheet.cell(row = i+1, column = 3).value
    cost_se = costsheet.cell(row = i+1, column = 4).value
    CostDict[cost_name] = (cost_type, cost_mean, cost_se)
del(CostDict['Parameter'])
#del(CostDict['None'])

#############################################################################################
############################################################################################

ResourceList = []
EventsList = []
UtilitiesList = []
QALY = []

with open('estimates.pickle', 'rb') as f:
    estimates = pickle.load(f)

with open('regcoeffs.pickle', 'rb') as f:
    regcoeffs = pickle.load(f)

from pprint import pprint

resources = []
events = []
natHist = []
QALYList = []

from Glb_CreateEntity import Entity
from Glb_ApplyInit import ApplyInit
from Glb_CancerFlags import CancerFlags


entity = Entity()
applyinit = ApplyInit(estimates)
applyinit.Process(entity)
entity.probOPL = 1      # All entities start with OPL
from NatHist_OralCancer import NatHistOCa
nathistoca = NatHistOCa(estimates, regcoeffs)
nathistoca.Process(entity, natHist)

#PROBABILITY NODE: Does this person regularly see a dentist?
# If yes - develop OPL while undergoing regular observations (state 1.0)
# If no - develop OPL and possibly cancer (state 1.8)

# In this scenario, all entities start at OPL management    
entity.OPLDetected = 1
entity.stateNum = 2.0
entity.altscen = 1
entity.hasDentist = 1


# Load a pickled entity instead of drawing a new one:
    # Dead of undetected stage IV:
"""
with open('entity_undetIV.pickle', 'rb') as f:
    entity = pickle.load(f)
    entity.time_Sysp = 0
    natHist.append(entity.natHist)
    natHistAr = numpy.asarray(entity.natHist)                     
    numpy.save('natHistAr', natHistAr)                            
"""
# Spontaneously resolving OPL:
"""
with open('entity_NED.pickle', 'rb') as f:
    entity = pickle.load(f)
    entity.time_Sysp = 0
    natHist.append(entity.natHist)
    natHistAr = numpy.asarray(entity.natHist)                     
    numpy.save('natHistAr', natHistAr)                            
"""
# Symptomatic detection:
"""
with open('entity_sympt.pickle', 'rb') as f:
    entity = pickle.load(f)
    entity.time_Sysp = 0
    natHist.append(entity.natHist)
    natHistAr = numpy.asarray(entity.natHist)                     
    numpy.save('natHistAr', natHistAr)                            
"""
# Detected and treated cancer:
"""    
with open('entity_hascancer.pickle', 'rb') as f:
    entity = pickle.load(f)
"""      

            
if entity.stateNum == 0.1:
    if entity.hasDentist == 1:
        entity.stateNum = 1.0
        entity.currentState = "1.0 - Undergoing regular dental screening"
    elif entity.hasDentist == 0:
        entity.stateNum = 1.8
        entity.currentState = "1.8 - No access to dentist"
    else:
        print("The entity has not been assigned a value for 'hasDentist'. The simulation must terminate")
        entity.stateNum = 0.9
    
### Advance the clock to next scheduled event (NatHist, Sysp, Recurrence, Death) ###

from Glb_CheckTime import CheckTime
CheckTime(entity, estimates, natHist, QALY)

### Run next scheduled event/process according to state ###

#People with a participating dentist undergo regular screening appointments    
if entity.stateNum == 1.0:                
    from SysP_ScreenAppt import ScreenAppt
    screenappt = ScreenAppt(estimates, regcoeffs)            
    screenappt.Process(entity)
   
#People with no dentist wait for disease event        
if entity.stateNum == 1.8:
    entity.time_Sysp += 1000      #Move system process clock forward by 1000 days (See footnote 1)
        
#People with a detected premalignancy undergo regular follow-up        
if entity.stateNum == 2.0:
    from Chap5_SysP_OPLmanage import OPLManage
    oplmanage = OPLManage(estimates, regcoeffs)            
    oplmanage.Process(entity) 

#People with a detected cancer undergo treatment       
if entity.stateNum == 3.0:
    from SysP_IncidentCancer import IncidentCancer
    incidentcancer = IncidentCancer(estimates, regcoeffs)
    incidentcancer.Process(entity)        

#People who have been successfully treated undergo regular follow-up     
if entity.stateNum == 4.0:
    from SysP_Followup import Followup
    followup = Followup(estimates, regcoeffs)
    followup.Process(entity)

#People whose disease has entered remission after 10 years     
if entity.stateNum == 4.8:
    #entity is in remission, no further events occur
    entity.allTime = entity.natHist_deathAge + 0.0001
     
#People with terminal disease receive palliative care     
if entity.stateNum == 5.0:
    from SysP_Terminal import Terminal
    terminal = Terminal(estimates, regcoeffs)
    terminal.Process(entity)

#The entity is dead      
if entity.stateNum == 100:
    #print("Entity is", entity.death_desc, "at:", entity.time_death)
    events.append(('Entity dies', entity.time_death))
    
# An error has occurred
if entity.stateNum == 99:
    print("An error has occurred and the simulation must end")
    print(entity.currentState)
    
    
pprint(vars(entity))
"""
ResourceList.append(resources)
EventsList.append(events)
QALYList.append(QALY)
"""
"""
numpy.save('ResourceList', ResourceList)
numpy.save('EventsList', EventsList)
numpy.save('QALYList', QALYList)

from NatHist_OralCancer import NatHistOCa
nathist = NatHistOCa(estimates, regcoeffs)
nathist.Process(entity, natHist)

entity.nh_status = 1.0
entity.OPLStatus = 1
entity.nh_status = 1.0
entity.natHist = []
entity.age = entity.startAge
entity.hasCancer = 1
entity.cancerStage = 'I'
entity.stateNum = 1.0
entity.utility = ("Undetected Stage I", estimates.Util_StageI_Undetected.sample(), entity.allTime)
"""